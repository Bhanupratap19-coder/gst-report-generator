import pandas as pd 
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# =======================================================
#       GST REPORT GENERATOR - PORTFOLIO PROJECT 
#       Made by: Bhanu Pratap 
# =======================================================

print("=" * 55)
print("     GST MONTHLY REPORT GENERATOR")
print("=" * 55)

# =======================================================
# STEP 1 - RAW DATA BANAO (CA FIRM KI FILE)
# =======================================================
raw_data = {
    "Client": ["Sharma & Co", "Gupta Traders", "Patel Ltd", "Verma GST",
               "Mehta Bros", "Singh & Co", "Joshi Ltd", "Kumar Traders",
               "Agarwal Co", "Tiwari GST", "Bansal Ltd", "Chopra & Co"],
    "GST_Amount": [12000, None, 23000, 5600,
                   18000, 9200, 31000, None,
                   15000, 11500, 27000, None],
    "Status": ["Paid", "Pending", "Paid", "Pending",
               "Paid", None, "Pending", "Paid",
               "Pending", "Paid", "Paid", "Pending"],
    "Month": ["March", "March", "April", "April",
              "March", "April", "March", "April",
              None, "April", "March", "April"]                                    
}

df = pd.DataFrame(raw_data)
print(f"\n Data load ho gaya - Total entries: {len(df)}")

#_________________________________________________________________
#  STEP 2 - NULL VALUES CLEAN KARO 
# ________________________________________________________________

print("\n Null values before cleaning:")
print(df.isnull().sum())

avg_gst = df["GST_Amount"].mean()
df["GST_Amount"] = df["GST_Amount"].fillna(round(avg_gst, 2))
df["Status"]     = df["Status"].fillna("Unknown")
df["Month"]      = df["Month"].fillna("Unknown")

print("\n Data clean ho gaya!")

# _________________________________________________________________
# STEP 3 - REPORTS BANAO 
# _________________________________________________________________

# Month wise summary 
month_report = df.groupby("Month")["GST_Amount"].agg(
    Total="sum",
    Average="mean",
    Max="max",
    Min="min",
    Clients="count"
).round(2)

# PAID CLIENTS 
paid = df[df["Status"] == "Paid"].sort_values(
    "GST_Amount", ascending=False).reset_index(drop=True)

# PENDING CLIENTS 
pending = df[df["Status"] == "Pending"].sort_values(
    "GST_Amount", ascending=False).reset_index(drop=True)


# ____________________________________________________________________
# STEP 4 - FINAL SUMMARY PRINT KARO 
# ____________________________________________________________________
print("\n" + "=" * 55)
print("            FINAL SUMMARY")
print("=" * 55)
print(f"  Total Clients    : {len(df)}")
print(f"  Paid  Clients    : {len(paid)}")
print(f"  Pending Clients  : {len(pending)}")
print(f"  Total GST        : ₹{df['GST_Amount'].sum():,.2f}")
print(f"  Collected GST    : ₹{paid['GST_Amount'].sum():,.2f}")
print(f"  Pending GST      : ₹{pending['GST_Amount'].sum():,.2f}")
print("=" * 55)

# ____________________________________________________________________
# STEP 5 - Excel mein Save karo (3 Sheets)
# ____________________________________________________________________
output_file = "GST_Final_Report.xlsx"

with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Full Data", index=False)
    month_report.to_excel(writer, sheet_name="Month Report")
    pending.to_excel(writer, sheet_name="Pending Clients", index=False)

# ____________________________________________________________________
# STEP 6 - PROFESSIONAL FORMATTING
# ____________________________________________________________________
wb = load_workbook(output_file)

# Colors define karo 
blue_fill  = PatternFill(start_color="FF2F75B6", end_color="FF2F75B6", fill_type="solid")
green_fill = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
red_fill   = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
white_font = Font(bold=True, color="FFFFFFFF", size=12)
bold_font  = Font(bold=True, size=11)

def format_header(ws):
    """Har sheet ka header blue + bold banao"""
    for cell in ws[1]:
        cell.fill = blue_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")

def fix_column_width(ws):
    """Column width auto fix karo""" 
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 4

def color_status_rows(ws, status_col_index):
    """Paid = Green, Pending = Red"""
    for row in ws.iter_rows(min_row=2):
        status_cell = row[status_col_index]
        if status_cell.value == "Paid":
            fill = green_fill
        elif status_cell.value == "Pending":
            fill = red_fill
        else:
            continue
        for cell in row:
            cell.fill = fill 

# ---- Sheet 1 - Full Data ----
ws1 = wb["Full Data"]
format_header(ws1)
color_status_rows(ws1, status_col_index=2) # C column = status
fix_column_width(ws1)

# ---- Sheet 2 -- Month Report ---
ws2 = wb["Month Report"]
format_header(ws2)
fix_column_width(ws2)
for cell in ws2[1]:
    cell.font = white_font

# ----Sheet 3 -- Pending Clients ----
ws3 = wb["Pending Clients"] 
format_header(ws3)
fix_column_width(ws3)
for row in ws3.iter_rows(min_row=2):
    for cell in row:
        cell.fill = red_fill 

#___________________________________________________________________
# STEP - FINAL SAVE 
# __________________________________________________________________

wb.save(output_file)

print(f"\n Professional report save ho gayi - {output_file}")
print("    Sheet 1 -- Full Data   ")
print("    Sheet 2 -- Month Report ")
print("    Sheet 3 -- Pending Clients")
print("\n Portfolio Project Complete!")
